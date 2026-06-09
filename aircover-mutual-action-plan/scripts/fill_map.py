#!/usr/bin/env python3
"""
fill_map.py  -  Fill a Mutual Action Plan template from a values file.

Template-agnostic: all coordinates, colours, status styles, and the sheet name
come from config.json, so a new customer can swap in their own MAP template and
declare its layout once without touching this code.

Usage:
    python fill_map.py --template <template.xlsx> --values values.json --out OUTPUT.xlsx [--config config.json]

If --config is omitted, the script looks for config.json next to the template,
then next to this script. If no config is found it falls back to the built-in
default layout (the shipped Aircover template).

What it does:
  - Writes ONLY the cells you supply. Everything else (banding, headers, borders,
    the =$F$7-Fn due-date formulas, the Status dropdown, merged ranges, widths) is
    left exactly as the template has it.
  - Overview fields render sourced values in black and unsourced values in grey
    italic (the house style: the rep sees what still needs their input).
  - The go-live field is written as a real date, so the template's due-date
    formulas recalculate from it.
  - When a Status changes, the matching colour from config is reapplied (the
    template hardcodes status colours per cell; there is no conditional format).

values.json format (every field optional; omit what you do not want to touch):
{
  "overview": {
    "account":       {"value": "Acme Corp",          "sourced": true},
    "project":       {"value": "Acme 2026",           "sourced": true},
    "objective":     {"value": "Cut ramp time 30%",   "sourced": true},
    "our_sponsor":   {"value": "Alex Rivera",        "sourced": true},
    "their_sponsor": {"value": "Jane Doe, VP Sales",  "sourced": false},
    "go_live":       {"value": "2026-10-01",          "sourced": true}
  },
  "action_items": [
    {"row": 12, "status": "Complete", "owner_us": "AE / SC", "owner_them": "VP Sales", "notes": "..."}
  ]
}

Requires openpyxl (pip install openpyxl --break-system-packages if missing).
"""
import argparse
import datetime
import json
import os
import sys

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl --break-system-packages")

# Built-in default layout (the shipped Aircover template). Used only if no
# config.json is found, so the script still runs standalone.
DEFAULT_CONFIG = {
    "template": {"sheet": "Mutual Action Plan"},
    "overview_cells": {
        "account": "B5", "project": "B6", "objective": "B7",
        "our_sponsor": "F5", "their_sponsor": "F6", "go_live": "F7",
    },
    "go_live_field": "go_live",
    "item_cols": {
        "action": "C", "owner_us": "D", "owner_them": "E",
        "days_before": "F", "status": "H", "notes": "I",
    },
    "first_item_row": 12,
    "last_item_row": 28,
    "value_colors": {"sourced": "FF1F2937", "placeholder": "FF6B7280"},
    "status_styles": {
        "Complete":    {"font": "FF166534", "fill": "FFDCFCE7"},
        "In Progress": {"font": "FF854D0E", "fill": "FFFEF9C3"},
        "Not Started": {"font": "FF6B7280", "fill": "FFF3F4F6"},
        "At Risk":     {"font": "FF991B1B", "fill": "FFFEE2E2"},
        "Blocked":     {"font": "FF991B1B", "fill": "FFFEE2E2"},
    },
}


def load_config(explicit, template_path):
    candidates = []
    if explicit:
        candidates.append(explicit)
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(template_path)), "config.json"))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json"))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "config.json"))
    for path in candidates:
        if path and os.path.exists(path):
            with open(path, encoding="utf-8") as fh:
                print(f"  Config: {os.path.relpath(path)}")
                return json.load(fh)
    print("  Config: none found, using built-in default layout.")
    return DEFAULT_CONFIG


def restyle_font(cell, color=None, italic=None):
    """Keep the cell's existing font, change only colour / italic."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold,
        italic=f.italic if italic is None else italic,
        color=color if color else (f.color.rgb if f.color else None),
    )


def parse_date(raw):
    if isinstance(raw, (datetime.datetime, datetime.date)):
        return datetime.datetime(raw.year, raw.month, raw.day)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%B %d, %Y"):
        try:
            return datetime.datetime.strptime(str(raw).strip(), fmt)
        except ValueError:
            continue
    return None


def write_overview(ws, overview, cfg):
    cells = cfg["overview_cells"]
    go_live_field = cfg.get("go_live_field", "go_live")
    sourced_c = cfg["value_colors"]["sourced"]
    placeholder_c = cfg["value_colors"]["placeholder"]
    filled, placeholders = [], []

    for key, cell_ref in cells.items():
        if key == go_live_field:
            continue
        field = overview.get(key)
        if not field or field.get("value") in (None, ""):
            continue
        cell = ws[cell_ref]
        cell.value = field["value"]
        if field.get("sourced", True):
            restyle_font(cell, color=sourced_c, italic=False)
            filled.append(key)
        else:
            restyle_font(cell, color=placeholder_c, italic=True)
            placeholders.append(key)

    gl = overview.get(go_live_field)
    if gl and gl.get("value") and go_live_field in cells:
        d = parse_date(gl["value"])
        if d is None:
            print(f"  WARNING: could not parse {go_live_field} {gl['value']!r}; left unchanged.")
        else:
            ws[cells[go_live_field]].value = d
            filled.append(go_live_field)
    return filled, placeholders


def write_items(ws, items, cfg):
    item_cols = cfg["item_cols"]
    lo, hi = cfg["first_item_row"], cfg["last_item_row"]
    status_styles = cfg["status_styles"]
    touched = []
    for item in items:
        row = item.get("row")
        if not isinstance(row, int) or not (lo <= row <= hi):
            print(f"  WARNING: skipping item with out-of-range row {row!r} (valid {lo}-{hi}).")
            continue
        for key, col in item_cols.items():
            if key not in item or item[key] is None:
                continue
            cell = ws[f"{col}{row}"]
            cell.value = item[key]
            if key == "status":
                style = status_styles.get(str(item[key]).strip())
                if style:
                    restyle_font(cell, color=style["font"], italic=False)
                    cell.fill = PatternFill("solid", fgColor=style["fill"])
                else:
                    print(f"  WARNING: unknown status {item[key]!r}; value set, colour left as-is. "
                          f"Valid: {', '.join(status_styles)}.")
        touched.append(row)
    return touched


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--values", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--config", default=None)
    args = ap.parse_args()

    cfg = load_config(args.config, args.template)

    with open(args.values, encoding="utf-8") as fh:
        values = json.load(fh)

    wb = load_workbook(args.template)
    sheet = cfg.get("template", {}).get("sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    filled, placeholders = write_overview(ws, values.get("overview", {}), cfg)
    touched = write_items(ws, values.get("action_items", []), cfg)

    wb.save(args.out)

    print(f"Saved {args.out}")
    if filled:
        print(f"  Overview filled (black):      {', '.join(filled)}")
    if placeholders:
        print(f"  Overview placeholders (grey): {', '.join(placeholders)}")
    if touched:
        print(f"  Action rows updated:          {', '.join(str(r) for r in touched)}")
    print("  Reminder: recalc the output so the Due Date column recalculates from the new Go-Live date.")


if __name__ == "__main__":
    main()
