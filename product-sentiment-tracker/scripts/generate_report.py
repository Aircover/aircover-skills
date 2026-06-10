#!/usr/bin/env python3
"""
generate_report.py
Reads a JSONL file of per-meeting agent results and produces an .xlsx report.
Raw Data sheet is the source of truth; all other sheets aggregate from it.
Usage: python3 generate_report.py <results.jsonl> <output.xlsx>
Requires: openpyxl
"""

import json, sys, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) < 3:
    print("Usage: python3 generate_report.py <results.jsonl> <output.xlsx> [agent_label]")
    sys.exit(1)

agent_label = sys.argv[3] if len(sys.argv) > 3 else "Product sentiment / VoC agent (selected at run time)"

# ── Read JSONL ──
meetings = []
with open(sys.argv[1]) as f:
    for line in f:
        line = line.strip()
        if line:
            meetings.append(json.loads(line))

print(f"Loaded {len(meetings)} meetings from JSONL")

# ── Helpers ──
def is_empty(val):
    return not val or str(val).strip().lower() in ("not found", "not found.", "")

def parse_signals(text, signal_type):
    """Split a multi-signal agent result into individual signal strings."""
    if is_empty(text):
        return []
    blocks = re.split(r'\n\n+', text.strip())
    signals = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        first_line = block.split('\n')[0].strip()
        if first_line:
            signals.append(first_line[:200])
    return signals if signals else [text.split('\n')[0][:200]]

# ── Styling ──
TEAL = "1A7A7A"
DARK_NAVY = "1B2A4A"
LIGHT_GRAY = "F5F5F5"
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor=TEAL)
title_font = Font(name="Georgia", bold=True, color=DARK_NAVY, size=14)
subtitle_font = Font(name="Calibri", italic=True, color="666666", size=10)
data_font = Font(name="Calibri", size=9, color="333333")
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
)
alt_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font; cell.alignment = wrap; cell.border = thin_border
        if alt:
            cell.fill = alt_fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Aggregate by account ──
accts = defaultdict(lambda: {"meetings": [], "pr": 0, "pg": 0, "fs": 0})
for m in meetings:
    a = accts[m["acct"]]
    a["meetings"].append(m)
    if not is_empty(m.get("pr", "")): a["pr"] += 1
    if not is_empty(m.get("pg", "")): a["pg"] += 1
    if not is_empty(m.get("fs", "")): a["fs"] += 1

account_rows = []
for name, data in accts.items():
    mtgs = data["meetings"]
    total = data["pr"] + data["pg"] + data["fs"]
    dates = sorted(m["date"] for m in mtgs)

    # Top product request (highest pr_c)
    pr_mtgs = [m for m in mtgs if not is_empty(m.get("pr", ""))]
    top_pr = "Not found"
    if pr_mtgs:
        best = max(pr_mtgs, key=lambda m: m.get("pr_c", 0))
        top_pr = best["pr"].split("\n")[0][:80] + f" (Conf: {best.get('pr_c', 0)})"

    # Top product gap (highest pg_c)
    pg_mtgs = [m for m in mtgs if not is_empty(m.get("pg", ""))]
    top_pg = "Not found"
    if pg_mtgs:
        best = max(pg_mtgs, key=lambda m: m.get("pg_c", 0))
        top_pg = best["pg"].split("\n")[0][:80] + f" (Conf: {best.get('pg_c', 0)})"

    # Top signal overall
    all_scored = [(m, m.get("pr_c", 0)) for m in pr_mtgs] + [(m, m.get("pg_c", 0)) for m in pg_mtgs]
    top_signal = "Not found"
    if all_scored:
        best_m, _ = max(all_scored, key=lambda x: x[1])
        src = best_m["pr"] if not is_empty(best_m.get("pr", "")) else best_m.get("pg", "")
        top_signal = src.split("\n")[0][:120]

    # Latest VoC
    latest_voc = ""
    for m in sorted(mtgs, key=lambda m: m["date"], reverse=True):
        if not is_empty(m.get("voc", "")):
            latest_voc = m["voc"][:250]
            break

    account_rows.append({
        "name": name, "mtg": len(mtgs), "signals": total,
        "pr": data["pr"], "pg": data["pg"], "fs": data["fs"],
        "top_pr": top_pr, "top_pg": top_pg, "top_signal": top_signal,
        "first": dates[0], "last": dates[-1],
        "blockers": "Yes" if data["pg"] >= 2 else "No",
        "latest_voc": latest_voc,
    })

account_rows.sort(key=lambda a: -a["signals"])

# ── Parse individual signals ──
signal_rows = []
for m in sorted(meetings, key=lambda x: (x["acct"], x["date"])):
    for field, stype, conf_key in [
        ("pr", "Product Request", "pr_c"), ("pg", "Product Gap", "pg_c"),
        ("fs", "Feature Adoption", None), ("ci", "Competitive Intel", None),
        ("ir", "Integration Request", "ir_c"),
    ]:
        val = m.get(field, "")
        if is_empty(val):
            continue
        parsed = parse_signals(val, stype)
        conf = m.get(conf_key, "") if conf_key else ""
        for sig_text in parsed:
            signal_rows.append({
                "acct": m["acct"], "date": m["date"], "type": stype,
                "signal": sig_text, "conf": conf, "url": m.get("url", ""),
            })

# ── Subtitle ──
subtitle_text = (
    f"{len(meetings)} meetings across {len(accts)} accounts, "
    f"{min(m['date'] for m in meetings)} to {max(m['date'] for m in meetings)}. "
    f"{len(signal_rows)} individual signals extracted."
)

# ═══════════════════════════════════════
# Build workbook
# ═══════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Sentiment Highlights ──
ws1 = wb.active; ws1.title = "Sentiment Highlights"
ws1.merge_cells("A1:H1"); ws1["A1"] = "Customer Sentiment Highlights"; ws1["A1"].font = title_font
ws1.merge_cells("A2:H2"); ws1["A2"] = subtitle_text; ws1["A2"].font = subtitle_font
h1 = ["Account", "Meetings", "Total Signals", "Product Requests", "Product Gaps",
      "Feature Signals", "Top Product Request", "Top Product Gap"]
w1 = [18, 10, 12, 14, 12, 12, 42, 42]
for i, h in enumerate(h1, 1): ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, len(h1))
for idx, a in enumerate(account_rows):
    r = 4 + idx
    vals = [a["name"], a["mtg"], a["signals"], a["pr"], a["pg"], a["fs"], a["top_pr"], a["top_pg"]]
    for c, v in enumerate(vals, 1): ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, len(h1), idx % 2 == 1)
set_widths(ws1, w1)
ws1.auto_filter.ref = f"A3:H{3 + len(account_rows)}"
ws1.freeze_panes = "A4"

# ── Sheet 2: Account Summary ──
ws2 = wb.create_sheet("Account Summary")
ws2.merge_cells("A1:K1"); ws2["A1"] = "Customer Sentiment by Account"; ws2["A1"].font = title_font
ws2.merge_cells("A2:K2"); ws2["A2"] = subtitle_text; ws2["A2"].font = subtitle_font
h2 = ["Account", "Meetings", "Signals", "PR", "PG", "FS", "Top Signal",
      "First", "Latest", "Blockers?", "VoC Summary (Latest)"]
w2 = [18, 10, 10, 8, 8, 8, 48, 12, 12, 10, 55]
for i, h in enumerate(h2, 1): ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(h2))
for idx, a in enumerate(account_rows):
    r = 4 + idx
    vals = [a["name"], a["mtg"], a["signals"], a["pr"], a["pg"], a["fs"],
            a["top_signal"], a["first"], a["last"], a["blockers"], a["latest_voc"]]
    for c, v in enumerate(vals, 1): ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, len(h2), idx % 2 == 1)
set_widths(ws2, w2)
ws2.auto_filter.ref = f"A3:K{3 + len(account_rows)}"
ws2.freeze_panes = "A4"

# ── Sheet 3: Signal Detail ──
ws3 = wb.create_sheet("Signal Detail")
h3 = ["Account", "Date", "Signal Type", "Signal", "Confidence", "Meeting URL"]
w3 = [18, 12, 16, 65, 10, 48]
for i, h in enumerate(h3, 1): ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(h3))
for idx, s in enumerate(signal_rows):
    r = 2 + idx
    vals = [s["acct"], s["date"], s["type"], s["signal"], s["conf"], s["url"]]
    for c, v in enumerate(vals, 1): ws3.cell(row=r, column=c, value=v)
    style_row(ws3, r, len(h3), idx % 2 == 1)
set_widths(ws3, w3)
ws3.auto_filter.ref = f"A1:F{1 + len(signal_rows)}"
ws3.freeze_panes = "A2"

# ── Sheet 4: Sentiment Summary ──
ws4 = wb.create_sheet("Sentiment Summary")
ws4.merge_cells("A1:D1")
ws4["A1"] = "Signal Type Distribution"
ws4["A1"].font = title_font
h4a = ["Signal Type", "Accounts", "Total Signals", "% of Signals"]
for i, h in enumerate(h4a, 1): ws4.cell(row=2, column=i, value=h)
style_header(ws4, 2, len(h4a))

total_pr = sum(a["pr"] for a in account_rows)
total_pg = sum(a["pg"] for a in account_rows)
total_fs = sum(a["fs"] for a in account_rows)
grand = total_pr + total_pg + total_fs or 1
dist = [
    ("Product Requests", sum(1 for a in account_rows if a["pr"] > 0), total_pr, f"{100 * total_pr // grand}%"),
    ("Product Gaps", sum(1 for a in account_rows if a["pg"] > 0), total_pg, f"{100 * total_pg // grand}%"),
    ("Feature Adoption", sum(1 for a in account_rows if a["fs"] > 0), total_fs, f"{100 * total_fs // grand}%"),
]
for idx, (st, ac, ts, pct) in enumerate(dist):
    r = 3 + idx
    for c, v in enumerate([st, ac, ts, pct], 1): ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, len(h4a), idx % 2 == 1)

ts = 8
ws4.merge_cells(f"A{ts}:D{ts}")
ws4[f"A{ts}"] = "Top Signal Sources by Account"
ws4[f"A{ts}"].font = Font(name="Georgia", bold=True, color=TEAL, size=12)
h4b = ["Account", "Product Requests", "Product Gaps", "Feature Signals"]
for i, h in enumerate(h4b, 1): ws4.cell(row=ts + 1, column=i, value=h)
style_header(ws4, ts + 1, len(h4b))
for idx, a in enumerate(account_rows):
    r = ts + 2 + idx
    for c, v in enumerate([a["name"], a["pr"], a["pg"], a["fs"]], 1):
        ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, len(h4b), idx % 2 == 1)
set_widths(ws4, [22, 16, 16, 16])

# ── Sheet 5: Raw Data (source of truth) ──
ws5 = wb.create_sheet("Raw Data")
h5 = ["Meeting ID", "Date", "Account", "Meeting Title",
      "Product Requests", "PR Conf", "Product Gaps", "PG Conf",
      "Feature Adoption Signals", "Competitive Intelligence",
      "Integration Requests", "IR Conf", "VoC Summary", "Meeting URL"]
w5 = [35, 12, 18, 35, 55, 8, 55, 8, 55, 55, 55, 8, 55, 48]
for i, h in enumerate(h5, 1): ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(h5))
for idx, m in enumerate(sorted(meetings, key=lambda x: (x["acct"], x["date"]))):
    r = 2 + idx
    vals = [
        m.get("mid", ""), m.get("date", ""), m.get("acct", ""), m.get("title", ""),
        m.get("pr", ""), m.get("pr_c", ""), m.get("pg", ""), m.get("pg_c", ""),
        m.get("fs", ""), m.get("ci", ""), m.get("ir", ""), m.get("ir_c", ""),
        m.get("voc", ""), m.get("url", ""),
    ]
    for c, v in enumerate(vals, 1): ws5.cell(row=r, column=c, value=v)
    style_row(ws5, r, len(h5), idx % 2 == 1)
set_widths(ws5, w5)
ws5.auto_filter.ref = f"A1:{get_column_letter(len(h5))}{1 + len(meetings)}"
ws5.freeze_panes = "A2"

# ── Sheet 6: Agent Properties Guide ──
ws6 = wb.create_sheet("Agent Properties Guide")
ws6.merge_cells("A1:F1")
ws6["A1"] = f"Agent: {agent_label}"
ws6["A1"].font = title_font
ws6.merge_cells("A2:F2")
ws6["A2"] = "Properties discovered at runtime from Production. Join on title, not key."
ws6["A2"].font = subtitle_font
gh = ["Property #", "Title", "Type", "Max Score", "Scored?", "Description"]
for i, h in enumerate(gh, 1): ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, len(gh))
guide = [
    [0, "Product Requests", "Extraction", 5, "Yes", "Explicit feature requests from the customer."],
    [1, "Product Gaps", "Extraction", 5, "Yes", "Existing features that fall short of needs."],
    [2, "Feature Adoption Signals", "Extraction", 0, "No", "Customer reactions: Active Use, Interested, Not Adopted, Dropped."],
    [3, "Competitive Intelligence", "Extraction", 0, "No", "Competitor mentions, comparisons, sentiment."],
    [4, "Integration Requests", "Extraction", 5, "Yes", "Integration needs for connecting to other systems."],
    [5, "VoC Summary", "Extraction", 0, "No", "2-4 sentence executive summary of all signals."],
]
for idx, gd in enumerate(guide):
    r = 4 + idx
    for c, v in enumerate(gd, 1): ws6.cell(row=r, column=c, value=v)
    style_row(ws6, r, len(gh), idx % 2 == 1)
set_widths(ws6, [12, 24, 12, 10, 10, 55])

# ── Save ──
wb.save(sys.argv[2])
print(f"Report saved: {sys.argv[2]}")
print(f"  Accounts: {len(account_rows)}")
print(f"  Raw Data rows: {len(meetings)}")
print(f"  Signal Detail rows: {len(signal_rows)}")
